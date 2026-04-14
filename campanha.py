import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import date, datetime
import gspread
from google.oauth2.service_account import Credentials

# ── Config ───────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CampanhaOS · Gestão",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed",
)

COLUNAS = [
    "id", "tipo", "cod_result", "filiais", "nome_acao", "laboratorio", "ponto_focal",
    "data_inicio", "data_fim", "duracao_dias", "reposicao_por",
    "margem_AL", "margem_PE", "margem_SE", "observacoes", "criado_em",
]

COLUNAS_APURACAO = [
    "id_campanha",
    "venda_liq_geral", "venda_liq_AL", "venda_liq_PE", "venda_liq_SE",
    "margem_med_geral", "margem_med_AL", "margem_med_PE", "margem_med_SE",
    "unidades_geral", "unidades_AL", "unidades_PE", "unidades_SE",
    "positivacoes_geral", "positivacoes_AL", "positivacoes_PE", "positivacoes_SE",
    "valor_reposicao",
    "apurado_em",
]

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

C_BG     = "#F2F4F7"
C_YELLOW = "#FFCA08"
C_BLUE   = "#004898"
C_TEAL   = "#09A49B"
C_BORDER = "#09A49B"
C_DARK   = "#003370"

# ── Tema ─────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&display=swap');

html, body, [data-testid="stAppViewContainer"], [data-testid="stApp"] {{
    background-color: {C_BG} !important;
    color: {C_BLUE} !important;
    font-family: 'Syne', sans-serif !important;
}}
[data-testid="stHeader"] {{ background: {C_BG} !important; }}
h1, h2, h3 {{ font-family: 'Syne', sans-serif !important; letter-spacing: -0.02em; }}
h1 {{ font-weight: 800; font-size: 2rem !important; color: {C_BLUE} !important; }}
h2 {{ font-weight: 700; font-size: 1.25rem !important; color: {C_BLUE} !important; }}
h3 {{ font-weight: 700; color: {C_TEAL} !important; font-size: 0.8rem !important;
     text-transform: uppercase; letter-spacing: 0.1em; }}
[data-testid="stTextInput"] input,
[data-testid="stTextAreaRootElement"] textarea,
[data-testid="stDateInput"] input,
[data-testid="stNumberInput"] input {{
    background-color: #FFFFFF !important;
    border: 1.5px solid {C_TEAL} !important;
    border-radius: 6px !important;
    color: {C_BLUE} !important;
    font-family: 'Syne', sans-serif !important;
    font-size: 0.9rem !important;
}}
[data-testid="stTextInput"] input:focus,
[data-testid="stTextAreaRootElement"] textarea:focus {{
    border-color: {C_BLUE} !important;
    box-shadow: 0 0 0 2px rgba(0,72,152,0.15) !important;
}}
[data-testid="stSelectbox"] > div > div {{
    background-color: #FFFFFF !important;
    border: 1.5px solid {C_TEAL} !important;
    border-radius: 6px !important;
    color: {C_BLUE} !important;
}}
label, [data-testid="stWidgetLabel"] p {{
    color: {C_BLUE} !important; font-size: 0.78rem !important;
    font-weight: 700 !important; text-transform: uppercase !important;
    letter-spacing: 0.07em !important; font-family: 'Syne', sans-serif !important;
}}
[data-testid="stButton"] > button[kind="primary"] {{
    background: {C_BLUE} !important; color: #FFFFFF !important;
    border: none !important; border-radius: 6px !important;
    font-family: 'Syne', sans-serif !important; font-weight: 800 !important;
    font-size: 0.9rem !important; padding: 0.6rem 2rem !important;
    transition: opacity 0.2s !important;
}}
[data-testid="stButton"] > button[kind="primary"]:hover {{ opacity: 0.88 !important; }}
[data-testid="stButton"] > button[kind="secondary"] {{
    background: #FFFFFF !important; color: {C_BLUE} !important;
    border: 1.5px solid {C_TEAL} !important; border-radius: 6px !important;
    font-family: 'Syne', sans-serif !important; font-weight: 600 !important;
}}
[data-testid="stTabs"] [role="tablist"] {{ border-bottom: 2px solid #D1DCE8 !important; gap: 0 !important; }}
[data-testid="stTabs"] button[role="tab"] {{
    background: transparent !important; color: #94A3B8 !important;
    font-family: 'Syne', sans-serif !important; font-weight: 700 !important;
    font-size: 0.85rem !important; border: none !important;
    padding: 0.6rem 1.4rem !important; border-bottom: 3px solid transparent !important;
}}
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {{
    color: {C_BLUE} !important; border-bottom: 3px solid {C_YELLOW} !important;
}}
[data-testid="stDataFrame"] {{ border: 1.5px solid {C_TEAL} !important; border-radius: 8px !important; overflow: hidden; }}
hr {{ border-color: #D1DCE8 !important; }}
[data-testid="stRadio"] label {{ text-transform: none !important; font-size: 0.9rem !important; color: {C_BLUE} !important; }}
[data-testid="stMetric"] {{
    background: {C_BLUE} !important; border: none !important;
    border-radius: 8px !important; padding: 1rem !important;
}}
[data-testid="stMetricLabel"] {{ color: {C_YELLOW} !important; font-size: 0.72rem !important; font-weight: 700 !important; }}
[data-testid="stMetricValue"] {{ color: #FFFFFF !important; font-size: 1.6rem !important; font-weight: 800 !important; }}
[data-testid="stCheckbox"] label {{ text-transform: none !important; font-size: 0.9rem !important; color: {C_BLUE} !important; }}
[data-testid="stDownloadButton"] > button {{
    background: #FFFFFF !important; border: 1.5px solid {C_TEAL} !important;
    border-radius: 6px !important; color: {C_TEAL} !important;
    font-family: 'Syne', sans-serif !important; font-weight: 600 !important;
    font-size: 0.82rem !important; padding: 0.45rem 1rem !important;
    transition: all 0.2s !important;
}}
[data-testid="stDownloadButton"] > button:hover {{
    background: {C_YELLOW} !important; color: {C_BLUE} !important; border-color: {C_YELLOW} !important;
}}
[data-testid="stNotification"] {{ border-radius: 6px !important; }}
::-webkit-scrollbar {{ width: 6px; height: 6px; }}
::-webkit-scrollbar-track {{ background: {C_BG}; }}
::-webkit-scrollbar-thumb {{ background: {C_TEAL}; border-radius: 3px; }}
[data-testid="stExpander"] {{
    background: #FFFFFF !important; border: 1.5px solid {C_TEAL} !important;
    border-radius: 8px !important;
}}
[data-testid="stExpander"] summary {{ color: {C_BLUE} !important; font-weight: 700 !important; }}
</style>
""", unsafe_allow_html=True)


# ── Google Sheets ─────────────────────────────────────────────────────────────
@st.cache_resource
def conectar_sheets():
    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"], scopes=SCOPES)
    return gspread.authorize(creds)

def get_worksheet(nome="Campanhas"):
    client = conectar_sheets()
    sh = client.open_by_key(st.secrets["SPREADSHEET_ID"])
    try:
        ws = sh.worksheet(nome)
    except gspread.WorksheetNotFound:
        cols = COLUNAS if nome == "Campanhas" else COLUNAS_APURACAO
        ws = sh.add_worksheet(title=nome, rows=1000, cols=len(cols))
        ws.append_row(cols)
    return ws

@st.cache_data(ttl=30)
def carregar_dados() -> pd.DataFrame:
    try:
        ws = get_worksheet("Campanhas")
        dados = ws.get_all_records()
        if not dados:
            return pd.DataFrame(columns=COLUNAS)
        df = pd.DataFrame(dados, dtype=str)
        for col in COLUNAS:
            if col not in df.columns:
                df[col] = ""
        return df[COLUNAS]
    except Exception as e:
        st.error(f"Erro ao carregar dados: {e}")
        return pd.DataFrame(columns=COLUNAS)

@st.cache_data(ttl=30)
def carregar_apuracoes() -> pd.DataFrame:
    try:
        ws = get_worksheet("Apurações")
        dados = ws.get_all_records()
        if not dados:
            return pd.DataFrame(columns=COLUNAS_APURACAO)
        df = pd.DataFrame(dados, dtype=str)
        for col in COLUNAS_APURACAO:
            if col not in df.columns:
                df[col] = ""
        return df[COLUNAS_APURACAO]
    except Exception:
        return pd.DataFrame(columns=COLUNAS_APURACAO)

def salvar_linha(nova: dict):
    ws = get_worksheet("Campanhas")
    ws.append_row([nova.get(col, "") for col in COLUNAS])
    st.cache_data.clear()

def atualizar_linha(id_registro: str, dados: dict):
    ws = get_worksheet("Campanhas")
    cell = ws.find(id_registro)
    if cell:
        for ci, col in enumerate(COLUNAS, start=1):
            ws.update_cell(cell.row, ci, dados.get(col, ""))
    st.cache_data.clear()

def excluir_linha(id_registro: str):
    ws = get_worksheet("Campanhas")
    cell = ws.find(id_registro)
    if cell:
        ws.delete_rows(cell.row)
    st.cache_data.clear()

def salvar_apuracao(dados: dict):
    ws = get_worksheet("Apurações")
    # Verificar se já existe apuração para esse id_campanha e sobrescrever
    try:
        cell = ws.find(dados["id_campanha"])
        if cell:
            for ci, col in enumerate(COLUNAS_APURACAO, start=1):
                ws.update_cell(cell.row, ci, dados.get(col, ""))
        else:
            ws.append_row([dados.get(col, "") for col in COLUNAS_APURACAO])
    except Exception:
        ws.append_row([dados.get(col, "") for col in COLUNAS_APURACAO])
    st.cache_data.clear()

def proximo_id(df: pd.DataFrame) -> str:
    if df.empty:
        return "CAM-001"
    nums = []
    for i in df["id"].dropna().tolist():
        try:
            nums.append(int(str(i).replace("CAM-", "")))
        except Exception:
            pass
    return f"CAM-{(max(nums) + 1):03d}" if nums else "CAM-001"

def parse_date(val: str):
    try:
        return datetime.strptime(val, "%d/%m/%Y").date()
    except Exception:
        return date.today()

def status_campanha(data_fim_str: str) -> tuple:
    """Retorna (label, cor_bg, cor_texto) com base na data de fim."""
    try:
        dt = datetime.strptime(data_fim_str, "%d/%m/%Y").date()
        hoje = date.today()
        if dt >= hoje:
            return "🟢 Em andamento", "#DCFCE7", "#166534"
        else:
            return "🔴 Encerrada", "#FEE2E2", "#991B1B"
    except Exception:
        return "—", "#F1F5F9", "#64748B"


# ── Exportador Excel ──────────────────────────────────────────────────────────
def gerar_excel(df: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Campanhas & Ações"
    thin  = Side(style="thin",   color="09A49B")
    thick = Side(style="medium", color="004898")
    borda     = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    borda_hdr = Border(left=thick, right=thick, top=thick, bottom=thick)

    ws.merge_cells("A1:O1")
    ws["A1"] = "CampanhaOS  —  Relatório de Campanhas & Ações"
    ws["A1"].font      = Font(name="Arial", bold=True, size=15, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="004898")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:O2")
    ws["A2"] = (
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        f"   |   Total: {len(df)}"
        f"   |   Campanhas: {len(df[df['tipo']=='Campanha'])}"
        f"   |   Ações: {len(df[df['tipo']=='Ação'])}"
    )
    ws["A2"].font      = Font(name="Arial", size=9, color="FFFFFF", italic=True)
    ws["A2"].fill      = PatternFill("solid", fgColor="09A49B")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    colunas = [
        ("ID",             "id",            10),
        ("Tipo",           "tipo",          13),
        ("Cód. Result",    "cod_result",    13),
        ("Filiais",        "filiais",       14),
        ("Nome",           "nome_acao",     30),
        ("Laboratório",    "laboratorio",   22),
        ("Ponto Focal",    "ponto_focal",   20),
        ("Início",         "data_inicio",   13),
        ("Fim",            "data_fim",      13),
        ("Duração (dias)", "duracao_dias",  13),
        ("Reposição",      "reposicao_por", 13),
        ("Margem AL (%)",  "margem_AL",     13),
        ("Margem PE (%)",  "margem_PE",     13),
        ("Margem SE (%)",  "margem_SE",     13),
        ("Cadastrado em",  "criado_em",     18),
    ]
    for ci, (lbl, _, w) in enumerate(colunas, 1):
        cell = ws.cell(row=4, column=ci, value=lbl)
        cell.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="004898")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = borda_hdr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 26

    tipo_bg = {"Campanha": "004898", "Ação": "09A49B"}
    tipo_fg = {"Campanha": "FFCA08", "Ação": "FFCA08"}
    for ri, (_, row) in enumerate(df.iterrows(), 5):
        fill_bg = "EBF4FF" if ri % 2 == 0 else "FFFFFF"
        tipo_v  = str(row.get("tipo", ""))
        for ci, (_, campo, _) in enumerate(colunas, 1):
            val  = str(row.get(campo, "")) if pd.notna(row.get(campo, "")) else ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = borda
            if campo == "tipo" and tipo_v in tipo_bg:
                cell.fill = PatternFill("solid", fgColor=tipo_bg[tipo_v])
                cell.font = Font(name="Arial", size=9, bold=True, color=tipo_fg[tipo_v])
            else:
                cell.fill = PatternFill("solid", fgColor=fill_bg)
                cell.font = Font(name="Arial", size=9, color="004898")
        ws.row_dimensions[ri].height = 18

    tot = 5 + len(df)
    ws.merge_cells(f"A{tot}:C{tot}")
    ws[f"A{tot}"] = f"Total: {len(df)} registro(s)"
    ws[f"A{tot}"].font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws[f"A{tot}"].fill      = PatternFill("solid", fgColor="09A49B")
    ws[f"A{tot}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[tot].height = 18
    ws.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Exportador PDF ────────────────────────────────────────────────────────────
def gerar_pdf(df: pd.DataFrame) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    def hc(h):
        h = h.lstrip("#")
        return colors.Color(int(h[0:2],16)/255, int(h[2:4],16)/255, int(h[4:6],16)/255)

    HEAD_C = hc("004898"); TEAL_C = hc("09A49B"); YEL_C = hc("FFCA08")
    EVEN_C = hc("EBF4FF"); GRAY_C = hc("64748B"); TEXT_C = hc("004898")

    sWH  = ParagraphStyle("wh",  fontSize=7, fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_CENTER)
    sCEL = ParagraphStyle("cel", fontSize=7, fontName="Helvetica",      textColor=TEXT_C,       alignment=TA_CENTER, leading=9)
    sCMP = ParagraphStyle("cmp", fontSize=7, fontName="Helvetica-Bold", textColor=hc("004898"), alignment=TA_CENTER)
    sACO = ParagraphStyle("aco", fontSize=7, fontName="Helvetica-Bold", textColor=hc("09A49B"), alignment=TA_CENTER)
    sTIT = ParagraphStyle("tit", fontSize=16,fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_LEFT)
    sSUB = ParagraphStyle("sub", fontSize=8, fontName="Helvetica",      textColor=colors.white, alignment=TA_LEFT)
    sROD = ParagraphStyle("rod", fontSize=7, fontName="Helvetica-Oblique", textColor=GRAY_C,    alignment=TA_RIGHT)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=12*mm,  bottomMargin=12*mm)
    story = []

    hdr = Table([[
        Paragraph("CampanhaOS", sTIT),
        Paragraph(
            f"Relatório de Campanhas &amp; Ações &nbsp;|&nbsp; "
            f"Gerado em: <b>{datetime.now().strftime('%d/%m/%Y %H:%M')}</b> &nbsp;|&nbsp; "
            f"Total: <b>{len(df)}</b>", sSUB),
    ]], colWidths=[55*mm, None])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(-1,-1), HEAD_C),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0),(-1,-1), 8),
        ("RIGHTPADDING", (0,0),(-1,-1), 8),
        ("TOPPADDING",   (0,0),(-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
    ]))
    story.append(hdr)
    story.append(Spacer(1, 4*mm))

    labels = ["Tipo","Cód.Result","Filiais","Nome","Laboratório","Ponto Focal",
              "Início","Fim","Dias","Reposição","Mg AL","Mg PE","Mg SE","Observações"]
    keys   = ["tipo","cod_result","filiais","nome_acao","laboratorio","ponto_focal",
              "data_inicio","data_fim","duracao_dias","reposicao_por",
              "margem_AL","margem_PE","margem_SE","observacoes"]
    widths = [18,18,18,46,34,28,18,18,10,20,13,13,13,30]

    tdata = [[Paragraph(l, sWH) for l in labels]]
    for _, row in df.iterrows():
        linha = []
        for k in keys:
            v = str(row.get(k,"")) if pd.notna(row.get(k,"")) else ""
            if k == "tipo":
                linha.append(Paragraph(v, sCMP if v == "Campanha" else sACO))
            else:
                linha.append(Paragraph(v, sCEL))
        tdata.append(linha)

    tbl = Table(tdata, colWidths=[w*mm for w in widths], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  HEAD_C),
        ("TOPPADDING",    (0,0),(-1,0),  5),
        ("BOTTOMPADDING", (0,0),(-1,0),  5),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, EVEN_C]),
        ("TOPPADDING",    (0,1),(-1,-1), 3),
        ("BOTTOMPADDING", (0,1),(-1,-1), 3),
        ("GRID",          (0,0),(-1,-1), 0.4, hc("B8D4F0")),
        ("LINEBELOW",     (0,0),(-1,0),  1.5, YEL_C),
        ("BOX",           (0,0),(-1,-1), 1.5, HEAD_C),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width="100%", thickness=0.8, color=YEL_C))
    story.append(Spacer(1, 1.5*mm))
    story.append(Paragraph("CampanhaOS · Gestão de Campanhas & Ações", sROD))
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ── Header ────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div style="display:flex;align-items:center;gap:12px;margin-bottom:0.5rem">
  <div style="width:40px;height:40px;background:{C_BLUE};
              border-radius:10px;display:flex;align-items:center;justify-content:center;
              font-size:1.3rem;color:#FFCA08;font-weight:800">⚡</div>
  <div>
    <div style="font-size:1.5rem;font-weight:800;color:{C_BLUE};line-height:1">CampanhaOS</div>
    <div style="font-size:0.72rem;color:{C_TEAL};font-weight:700;letter-spacing:0.1em;text-transform:uppercase">
      Gestão de Campanhas & Ações
    </div>
  </div>
</div>
<hr style="margin:0.8rem 0 1.5rem;border-color:#D1DCE8"/>
""", unsafe_allow_html=True)

tab_novo, tab_lista, tab_editar, tab_apuracao = st.tabs([
    "＋  Nova Campanha / Ação",
    "📋  Registros",
    "✏️  Editar / Excluir",
    "📊  Apuração",
])


# ══ FORMULÁRIO REUTILIZÁVEL ═══════════════════════════════════════════════════
def render_form(key_prefix: str, defaults: dict = None):
    d = defaults or {}

    st.markdown("### Filiais participantes")
    filiais_default = [f.strip() for f in d.get("filiais","AL").split(",")]
    col_al, col_pe, col_se, _ = st.columns([1,1,1,4])
    with col_al: fal = st.checkbox("AL — Alagoas",    value="AL" in filiais_default, key=f"{key_prefix}_al")
    with col_pe: fpe = st.checkbox("PE — Pernambuco", value="PE" in filiais_default, key=f"{key_prefix}_pe")
    with col_se: fse = st.checkbox("SE — Sergipe",    value="SE" in filiais_default, key=f"{key_prefix}_se")
    filiais_sel = [f for f,v in [("AL",fal),("PE",fpe),("SE",fse)] if v]

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("### Tipo de registro")
    tipo_idx = 0 if d.get("tipo","Campanha") == "Campanha" else 1
    tipo = st.radio("Tipo", options=["Campanha","Ação"], horizontal=True,
                    index=tipo_idx, label_visibility="collapsed", key=f"{key_prefix}_tipo")

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("### Configurações")

    if tipo == "Ação":
        rep_idx = 0 if d.get("reposicao_por","Fornecedor") == "Fornecedor" else 1
        reposicao_por = st.selectbox("Reposição paga por", ["Fornecedor","Empresa"],
                                     index=rep_idx, key=f"{key_prefix}_rep")
        # Código Result — só para Ação
        cod_result = st.text_input(
            "Código da Ação (Result)",
            value=d.get("cod_result",""),
            placeholder="Ex: ACT-2025-001",
            key=f"{key_prefix}_cod",
        )
    else:
        reposicao_por = ""
        cod_result    = ""
        st.markdown(
            '<div style="padding:0.6rem 1rem;background:#EEF3FA;border:1.5px solid #09A49B;'
            'border-radius:6px;color:#64748B;font-size:0.82rem;display:inline-block">'
            'Reposição e margem não se aplicam para Campanha</div>',
            unsafe_allow_html=True,
        )

    exibir_margem = (tipo == "Ação") and (reposicao_por == "Fornecedor")
    margem_vals   = {}

    if exibir_margem:
        if not filiais_sel:
            st.markdown(
                '<div style="padding:0.75rem;background:#FEF2F2;border:1.5px solid #EF4444;'
                'border-radius:6px;color:#DC2626;font-size:0.82rem;margin-top:0.5rem">'
                '⚠ Selecione ao menos uma filial para preencher as margens.</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown('<div style="font-size:0.78rem;color:#64748B;margin:0.6rem 0">Margem média por filial</div>', unsafe_allow_html=True)
            mc = st.columns(len(filiais_sel))
            for idx, filial in enumerate(filiais_sel):
                with mc[idx]:
                    margem_vals[filial] = st.text_input(
                        f"Margem {filial} (%)",
                        value=d.get(f"margem_{filial}",""),
                        placeholder="Ex: 18,5",
                        key=f"{key_prefix}_mg_{filial}",
                    )
    elif tipo == "Ação" and reposicao_por == "Empresa":
        st.markdown(
            '<div style="padding:0.6rem 1rem;background:#EEF3FA;border:1.5px solid #09A49B;'
            'border-radius:6px;color:#64748B;font-size:0.82rem;display:inline-block;margin-top:0.5rem">'
            'Margem não aplicável quando reposição é pela Empresa</div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

    with st.form(f"form_{key_prefix}", clear_on_submit=(key_prefix=="novo")):
        st.markdown("### Identificação")
        c1, c2 = st.columns(2)
        with c1: nome_acao   = st.text_input("Nome da ação / campanha",
                                              value=d.get("nome_acao",""),
                                              placeholder="Ex: Verão 2025 — Linha Infantil")
        with c2: laboratorio = st.text_input("Laboratório / Fornecedor",
                                              value=d.get("laboratorio",""),
                                              placeholder="Ex: Grupo Hypermarcas")
        ponto_focal = st.text_input("Ponto focal (fornecedor)",
                                    value=d.get("ponto_focal",""),
                                    placeholder="Ex: João Silva")

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("### Período")
        d1c, d2c = st.columns(2)
        with d1c: data_inicio = st.date_input("Data de início", value=parse_date(d.get("data_inicio","")), format="DD/MM/YYYY", key=f"{key_prefix}_di")
        with d2c: data_fim    = st.date_input("Data de fim",    value=parse_date(d.get("data_fim","")),    format="DD/MM/YYYY", key=f"{key_prefix}_df")

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("### Observações")
        observacoes = st.text_area("Observações", value=d.get("observacoes",""),
                                   placeholder="Detalhes adicionais, metas...",
                                   height=90, label_visibility="collapsed")

        st.markdown("<br>", unsafe_allow_html=True)
        label_btn = "✓  Cadastrar registro" if key_prefix == "novo" else "💾  Salvar alterações"
        submitted = st.form_submit_button(label_btn, type="primary", use_container_width=True)

    return submitted, {
        "filiais_sel": filiais_sel, "tipo": tipo,
        "reposicao_por": reposicao_por, "cod_result": cod_result,
        "margem_vals": margem_vals, "nome_acao": nome_acao,
        "laboratorio": laboratorio, "ponto_focal": ponto_focal,
        "data_inicio": data_inicio, "data_fim": data_fim,
        "observacoes": observacoes,
    }


def montar_registro(fd: dict, id_val: str, criado_em: str = "") -> dict:
    return {
        "id":            id_val,
        "tipo":          fd["tipo"],
        "cod_result":    fd["cod_result"],
        "filiais":       ", ".join(fd["filiais_sel"]),
        "nome_acao":     fd["nome_acao"].strip(),
        "laboratorio":   fd["laboratorio"].strip(),
        "ponto_focal":   fd["ponto_focal"].strip(),
        "data_inicio":   fd["data_inicio"].strftime("%d/%m/%Y"),
        "data_fim":      fd["data_fim"].strftime("%d/%m/%Y"),
        "duracao_dias":  str((fd["data_fim"] - fd["data_inicio"]).days + 1),
        "reposicao_por": fd["reposicao_por"],
        "margem_AL":     fd["margem_vals"].get("AL","").strip(),
        "margem_PE":     fd["margem_vals"].get("PE","").strip(),
        "margem_SE":     fd["margem_vals"].get("SE","").strip(),
        "observacoes":   fd["observacoes"].strip(),
        "criado_em":     criado_em or datetime.now().strftime("%d/%m/%Y %H:%M"),
    }


# ══ TAB 1 — CADASTRO ══════════════════════════════════════════════════════════
with tab_novo:
    submitted, fd = render_form("novo")
    if submitted:
        if not fd["filiais_sel"]:
            st.error("Selecione ao menos uma filial.")
        elif not fd["nome_acao"].strip():
            st.error("Preencha o nome da ação / campanha.")
        elif not fd["laboratorio"].strip():
            st.error("Preencha o laboratório / fornecedor.")
        else:
            df_atual = carregar_dados()
            nova = montar_registro(fd, proximo_id(df_atual))
            salvar_linha(nova)
            st.success(f"✓ Registro **{nova['id']}** cadastrado com sucesso!")


# ══ TAB 2 — LISTAGEM ══════════════════════════════════════════════════════════
with tab_lista:
    df = carregar_dados()
    df_ap = carregar_apuracoes()

    if df.empty:
        st.markdown("""
        <div style="text-align:center;padding:3rem;color:#94A3B8">
          <div style="font-size:2rem;margin-bottom:0.5rem">📭</div>
          <div style="font-weight:700">Nenhum registro ainda</div>
          <div style="font-size:0.85rem;margin-top:0.25rem">Cadastre a primeira campanha na aba ao lado</div>
        </div>""", unsafe_allow_html=True)
    else:
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Total de registros",  len(df))
        m2.metric("Campanhas",           len(df[df["tipo"]=="Campanha"]))
        m3.metric("Ações",               len(df[df["tipo"]=="Ação"]))
        m4.metric("Fornecedores únicos", df["laboratorio"].nunique())

        st.markdown("<br>", unsafe_allow_html=True)

        cf1,cf2,cf3 = st.columns(3)
        with cf1: filtro_tipo   = st.selectbox("Tipo",    ["Todos","Campanha","Ação"])
        with cf2:
            labs = ["Todos"] + sorted(df["laboratorio"].dropna().unique().tolist())
            filtro_lab = st.selectbox("Laboratório", labs)
        with cf3: filtro_filial = st.selectbox("Filial",  ["Todas","AL","PE","SE"])

        cf4,cf5,cf6 = st.columns(3)
        with cf4:
            focais = ["Todos"] + sorted(df["ponto_focal"].dropna().replace("",pd.NA).dropna().unique().tolist())
            filtro_focal = st.selectbox("Ponto focal", focais)
        with cf5: filtro_status = st.selectbox("Status",  ["Todos","Em andamento","Encerrada"])
        with cf6: filtro_busca  = st.text_input("Buscar por nome", placeholder="Digite parte do nome...")

        df_view = df.copy()
        if filtro_tipo   != "Todos":  df_view = df_view[df_view["tipo"] == filtro_tipo]
        if filtro_lab    != "Todos":  df_view = df_view[df_view["laboratorio"] == filtro_lab]
        if filtro_filial != "Todas":  df_view = df_view[df_view["filiais"].str.contains(filtro_filial, na=False)]
        if filtro_focal  != "Todos":  df_view = df_view[df_view["ponto_focal"] == filtro_focal]
        if filtro_busca:              df_view = df_view[df_view["nome_acao"].str.contains(filtro_busca, case=False, na=False)]
        if filtro_status != "Todos":
            def chk(r):
                lbl, _, _ = status_campanha(r.get("data_fim",""))
                return "andamento" in lbl if filtro_status == "Em andamento" else "Encerrada" in lbl
            df_view = df_view[df_view.apply(chk, axis=1)]

        st.markdown(f"<div style='color:#64748B;font-size:0.8rem;margin-bottom:0.8rem'>{len(df_view)} registro(s)</div>", unsafe_allow_html=True)

        # IDs com apuração
        ids_apurados = set(df_ap["id_campanha"].tolist()) if not df_ap.empty else set()

        # Renderizar cards por registro
        for _, row in df_view.iterrows():
            lbl_status, bg_status, fg_status = status_campanha(row.get("data_fim",""))
            tem_apuracao = row["id"] in ids_apurados
            badge_ap = (f'<span style="background:#F0FDF4;color:#166534;border:1px solid #BBF7D0;'
                        f'padding:2px 8px;border-radius:4px;font-size:0.72rem;font-weight:700;margin-left:6px">✓ Apurada</span>'
                        if tem_apuracao else "")

            with st.expander(
                f"{row['id']}  ·  {row['nome_acao']}  ·  {row['laboratorio']}",
                expanded=False,
            ):
                r1c1, r1c2, r1c3, r1c4 = st.columns(4)
                with r1c1:
                    st.markdown(f"""
                    <div style="font-size:0.72rem;color:#64748B;font-weight:700;text-transform:uppercase;letter-spacing:0.05em">Tipo</div>
                    <div style="font-size:0.95rem;font-weight:800;color:{C_BLUE}">{row.get('tipo','—')}</div>
                    """, unsafe_allow_html=True)
                with r1c2:
                    st.markdown(f"""
                    <div style="font-size:0.72rem;color:#64748B;font-weight:700;text-transform:uppercase;letter-spacing:0.05em">Filiais</div>
                    <div style="font-size:0.95rem;font-weight:700;color:{C_TEAL}">{row.get('filiais','—')}</div>
                    """, unsafe_allow_html=True)
                with r1c3:
                    st.markdown(f"""
                    <div style="font-size:0.72rem;color:#64748B;font-weight:700;text-transform:uppercase;letter-spacing:0.05em">Período</div>
                    <div style="font-size:0.85rem;font-weight:700;color:{C_BLUE}">{row.get('data_inicio','—')} → {row.get('data_fim','—')}</div>
                    """, unsafe_allow_html=True)
                with r1c4:
                    st.markdown(f"""
                    <div style="font-size:0.72rem;color:#64748B;font-weight:700;text-transform:uppercase;letter-spacing:0.05em">Status</div>
                    <span style="background:{bg_status};color:{fg_status};padding:3px 10px;border-radius:4px;font-size:0.8rem;font-weight:700">{lbl_status}</span>
                    {badge_ap}
                    """, unsafe_allow_html=True)

                st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

                r2c1, r2c2, r2c3, r2c4 = st.columns(4)
                with r2c1:
                    st.markdown(f"<div style='font-size:0.72rem;color:#64748B;font-weight:700;text-transform:uppercase'>Ponto Focal</div><div style='font-size:0.88rem;color:{C_BLUE}'>{row.get('ponto_focal','—')}</div>", unsafe_allow_html=True)
                with r2c2:
                    st.markdown(f"<div style='font-size:0.72rem;color:#64748B;font-weight:700;text-transform:uppercase'>Reposição</div><div style='font-size:0.88rem;color:{C_BLUE}'>{row.get('reposicao_por','—') or '—'}</div>", unsafe_allow_html=True)
                with r2c3:
                    mg = " / ".join([f"{f}:{row.get(f'margem_{f}','—')}%" for f in ["AL","PE","SE"] if row.get(f"margem_{f}","")])
                    st.markdown(f"<div style='font-size:0.72rem;color:#64748B;font-weight:700;text-transform:uppercase'>Margens</div><div style='font-size:0.88rem;color:{C_BLUE}'>{mg or '—'}</div>", unsafe_allow_html=True)
                with r2c4:
                    cod = row.get("cod_result","") or "—"
                    st.markdown(f"<div style='font-size:0.72rem;color:#64748B;font-weight:700;text-transform:uppercase'>Cód. Result</div><div style='font-size:0.88rem;color:{C_BLUE}'>{cod}</div>", unsafe_allow_html=True)

                if row.get("observacoes",""):
                    st.markdown(f"<div style='margin-top:8px;padding:0.6rem 0.8rem;background:#EEF3FA;border-left:3px solid {C_TEAL};border-radius:4px;font-size:0.85rem;color:#334155'>{row['observacoes']}</div>", unsafe_allow_html=True)

                # Resumo da apuração se existir
                if tem_apuracao:
                    ap = df_ap[df_ap["id_campanha"] == row["id"]].iloc[0]
                    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                    st.markdown(f"<div style='font-size:0.72rem;font-weight:700;color:{C_TEAL};text-transform:uppercase;letter-spacing:0.06em;margin-bottom:4px'>Resumo da Apuração</div>", unsafe_allow_html=True)
                    ac1,ac2,ac3,ac4,ac5 = st.columns(5)
                    with ac1: st.markdown(f"<div style='font-size:0.7rem;color:#64748B;font-weight:700'>Venda Líq. Geral</div><div style='font-size:0.92rem;font-weight:800;color:{C_BLUE}'>R$ {ap.get('venda_liq_geral','—')}</div>", unsafe_allow_html=True)
                    with ac2: st.markdown(f"<div style='font-size:0.7rem;color:#64748B;font-weight:700'>Margem Geral</div><div style='font-size:0.92rem;font-weight:800;color:{C_BLUE}'>{ap.get('margem_med_geral','—')}%</div>", unsafe_allow_html=True)
                    with ac3: st.markdown(f"<div style='font-size:0.7rem;color:#64748B;font-weight:700'>Unidades</div><div style='font-size:0.92rem;font-weight:800;color:{C_BLUE}'>{ap.get('unidades_geral','—')}</div>", unsafe_allow_html=True)
                    with ac4: st.markdown(f"<div style='font-size:0.7rem;color:#64748B;font-weight:700'>Positivações</div><div style='font-size:0.92rem;font-weight:800;color:{C_BLUE}'>{ap.get('positivacoes_geral','—')}</div>", unsafe_allow_html=True)
                    with ac5: st.markdown(f"<div style='font-size:0.7rem;color:#64748B;font-weight:700'>Valor Reposição</div><div style='font-size:0.92rem;font-weight:800;color:{C_BLUE}'>R$ {ap.get('valor_reposicao','—')}</div>", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("### Exportar registros filtrados")

        exp1, exp2, exp3 = st.columns(3)
        ts = datetime.now().strftime("%d%m%Y_%H%M")
        card_style = f"background:#FFFFFF;border:1.5px solid {C_TEAL};border-radius:8px;padding:1rem 1.25rem;margin-bottom:0.5rem"

        with exp1:
            st.markdown(f'<div style="{card_style}"><div style="font-size:0.75rem;font-weight:700;color:{C_YELLOW};text-transform:uppercase;letter-spacing:0.06em;margin-bottom:0.3rem">📄 CSV</div><div style="font-size:0.8rem;color:#64748B">Formato simples, compatível com qualquer planilha</div></div>', unsafe_allow_html=True)
            st.download_button("⬇  Baixar CSV", data=df_view.to_csv(index=False).encode("utf-8"),
                               file_name=f"campanhas_{ts}.csv", mime="text/csv", use_container_width=True)
        with exp2:
            st.markdown(f'<div style="{card_style}"><div style="font-size:0.75rem;font-weight:700;color:{C_YELLOW};text-transform:uppercase;letter-spacing:0.06em;margin-bottom:0.3rem">📊 Excel (.xlsx)</div><div style="font-size:0.8rem;color:#64748B">Relatório formatado com cores, cabeçalho e totais</div></div>', unsafe_allow_html=True)
            st.download_button("⬇  Baixar Excel", data=gerar_excel(df_view),
                               file_name=f"campanhas_{ts}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        with exp3:
            st.markdown(f'<div style="{card_style}"><div style="font-size:0.75rem;font-weight:700;color:{C_YELLOW};text-transform:uppercase;letter-spacing:0.06em;margin-bottom:0.3rem">📑 PDF</div><div style="font-size:0.8rem;color:#64748B">Relatório pronto para impressão e compartilhamento</div></div>', unsafe_allow_html=True)
            st.download_button("⬇  Baixar PDF", data=gerar_pdf(df_view),
                               file_name=f"campanhas_{ts}.pdf", mime="application/pdf",
                               use_container_width=True)


# ══ TAB 3 — EDITAR / EXCLUIR ══════════════════════════════════════════════════
with tab_editar:
    df = carregar_dados()
    if df.empty:
        st.markdown('<div style="text-align:center;padding:3rem;color:#94A3B8"><div style="font-size:2rem">📭</div><div style="font-weight:700">Nenhum registro para editar</div></div>', unsafe_allow_html=True)
    else:
        opcoes = {f"{r['id']} — {r['nome_acao']}": r['id'] for _, r in df.iterrows()}
        escolha = st.selectbox("Selecione o registro", list(opcoes.keys()), key="sel_editar")
        id_sel  = opcoes[escolha]
        registro = df[df["id"] == id_sel].iloc[0].to_dict()

        st.markdown("<br>", unsafe_allow_html=True)
        col_ed, col_ex = st.columns([5, 1])

        with col_ed:
            st.markdown(f'<div style="font-size:0.75rem;font-weight:700;color:{C_BLUE};text-transform:uppercase;letter-spacing:0.08em;margin-bottom:0.8rem">Editando: {id_sel}</div>', unsafe_allow_html=True)
            submitted_ed, fd = render_form("editar", defaults=registro)
            if submitted_ed:
                if not fd["filiais_sel"]:
                    st.error("Selecione ao menos uma filial.")
                elif not fd["nome_acao"].strip():
                    st.error("Preencha o nome da ação / campanha.")
                elif not fd["laboratorio"].strip():
                    st.error("Preencha o laboratório / fornecedor.")
                else:
                    atualizado = montar_registro(fd, id_sel, registro.get("criado_em",""))
                    atualizar_linha(id_sel, atualizado)
                    st.success(f"✓ Registro **{id_sel}** atualizado com sucesso!")

        with col_ex:
            st.markdown('<div style="font-size:0.75rem;font-weight:700;color:#DC2626;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:0.8rem">Excluir</div>', unsafe_allow_html=True)
            st.markdown('<div style="font-size:0.8rem;color:#64748B;margin-bottom:1rem">Esta ação é permanente.</div>', unsafe_allow_html=True)
            if st.button("🗑  Excluir registro", key="btn_excluir"):
                st.session_state["confirmar_exclusao"] = id_sel
            if st.session_state.get("confirmar_exclusao") == id_sel:
                st.markdown(f'<div style="padding:0.75rem;background:#FEF2F2;border:1.5px solid #EF4444;border-radius:6px;color:#DC2626;font-size:0.82rem;margin-bottom:0.5rem">Confirmar exclusão de <b>{id_sel}</b>?</div>', unsafe_allow_html=True)
                cc1, cc2 = st.columns(2)
                with cc1:
                    if st.button("✓ Confirmar", key="confirmar_sim", type="primary"):
                        excluir_linha(id_sel)
                        st.session_state.pop("confirmar_exclusao", None)
                        st.success(f"Registro {id_sel} excluído.")
                        st.rerun()
                with cc2:
                    if st.button("✗ Cancelar", key="confirmar_nao"):
                        st.session_state.pop("confirmar_exclusao", None)
                        st.rerun()


# ══ TAB 4 — APURAÇÃO ══════════════════════════════════════════════════════════
with tab_apuracao:
    df = carregar_dados()
    df_ap = carregar_apuracoes()

    if df.empty:
        st.markdown('<div style="text-align:center;padding:3rem;color:#94A3B8"><div style="font-size:2rem">📭</div><div style="font-weight:700">Nenhuma campanha cadastrada ainda</div></div>', unsafe_allow_html=True)
    else:
        opcoes_ap = {f"{r['id']} — {r['nome_acao']} ({r['tipo']})": r['id'] for _, r in df.iterrows()}
        escolha_ap = st.selectbox("Selecione a campanha / ação para apurar", list(opcoes_ap.keys()), key="sel_apuracao")
        id_ap = opcoes_ap[escolha_ap]
        reg_ap = df[df["id"] == id_ap].iloc[0].to_dict()
        filiais_camp = [f.strip() for f in reg_ap.get("filiais","").split(",") if f.strip()]
        tem_reposicao = reg_ap.get("reposicao_por","") == "Fornecedor"

        # Defaults se já existe apuração
        ap_defaults = {}
        if not df_ap.empty and id_ap in df_ap["id_campanha"].values:
            ap_defaults = df_ap[df_ap["id_campanha"] == id_ap].iloc[0].to_dict()

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f"""
        <div style="background:#FFFFFF;border:1.5px solid {C_TEAL};border-radius:8px;padding:1rem 1.25rem;margin-bottom:1rem">
          <div style="font-size:0.72rem;color:#64748B;font-weight:700;text-transform:uppercase;letter-spacing:0.05em">Apurando</div>
          <div style="font-size:1.1rem;font-weight:800;color:{C_BLUE}">{reg_ap.get('nome_acao','')} <span style="font-size:0.8rem;font-weight:600;color:{C_TEAL}">· {reg_ap.get('laboratorio','')} · {reg_ap.get('data_inicio','')} → {reg_ap.get('data_fim','')}</span></div>
        </div>
        """, unsafe_allow_html=True)

        with st.form("form_apuracao", clear_on_submit=False):

            # ── Venda Líquida ─────────────────────────────────────────────────
            st.markdown("### Venda Líquida (R$)")
            vc = st.columns([2] + [1]*len(filiais_camp))
            with vc[0]: vl_geral = st.text_input("Geral", value=ap_defaults.get("venda_liq_geral",""), placeholder="Ex: 125000,00")
            filiais_vl = {}
            for i, fil in enumerate(filiais_camp):
                with vc[i+1]: filiais_vl[fil] = st.text_input(f"Filial {fil}", value=ap_defaults.get(f"venda_liq_{fil}",""), placeholder="Ex: 45000,00", key=f"vl_{fil}")

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Margem Média ──────────────────────────────────────────────────
            st.markdown("### Margem Média (%)")
            mc = st.columns([2] + [1]*len(filiais_camp))
            with mc[0]: mg_geral = st.text_input("Geral", value=ap_defaults.get("margem_med_geral",""), placeholder="Ex: 18,5", key="mg_g")
            filiais_mg = {}
            for i, fil in enumerate(filiais_camp):
                with mc[i+1]: filiais_mg[fil] = st.text_input(f"Filial {fil}", value=ap_defaults.get(f"margem_med_{fil}",""), placeholder="Ex: 18,5", key=f"mg_{fil}")

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Unidades Vendidas ─────────────────────────────────────────────
            st.markdown("### Total de Unidades Vendidas")
            uc = st.columns([2] + [1]*len(filiais_camp))
            with uc[0]: un_geral = st.text_input("Geral", value=ap_defaults.get("unidades_geral",""), placeholder="Ex: 3500", key="un_g")
            filiais_un = {}
            for i, fil in enumerate(filiais_camp):
                with uc[i+1]: filiais_un[fil] = st.text_input(f"Filial {fil}", value=ap_defaults.get(f"unidades_{fil}",""), placeholder="Ex: 1200", key=f"un_{fil}")

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Positivações ──────────────────────────────────────────────────
            st.markdown("### Positivações (clientes atendidos)")
            pc = st.columns([2] + [1]*len(filiais_camp))
            with pc[0]: pos_geral = st.text_input("Geral", value=ap_defaults.get("positivacoes_geral",""), placeholder="Ex: 280", key="pos_g")
            filiais_pos = {}
            for i, fil in enumerate(filiais_camp):
                with pc[i+1]: filiais_pos[fil] = st.text_input(f"Filial {fil}", value=ap_defaults.get(f"positivacoes_{fil}",""), placeholder="Ex: 95", key=f"pos_{fil}")

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Reposição (só se aplicar) ─────────────────────────────────────
            if tem_reposicao:
                st.markdown("### Valor Total de Reposição (R$)")
                val_rep = st.text_input("Valor de reposição", value=ap_defaults.get("valor_reposicao",""), placeholder="Ex: 8500,00", label_visibility="collapsed")
            else:
                val_rep = ""

            st.markdown("<br>", unsafe_allow_html=True)
            salvar_ap = st.form_submit_button("💾  Salvar Apuração", type="primary", use_container_width=True)

        if salvar_ap:
            ap_dados = {
                "id_campanha":        id_ap,
                "venda_liq_geral":    vl_geral.strip(),
                "venda_liq_AL":       filiais_vl.get("AL","").strip(),
                "venda_liq_PE":       filiais_vl.get("PE","").strip(),
                "venda_liq_SE":       filiais_vl.get("SE","").strip(),
                "margem_med_geral":   mg_geral.strip(),
                "margem_med_AL":      filiais_mg.get("AL","").strip(),
                "margem_med_PE":      filiais_mg.get("PE","").strip(),
                "margem_med_SE":      filiais_mg.get("SE","").strip(),
                "unidades_geral":     un_geral.strip(),
                "unidades_AL":        filiais_un.get("AL","").strip(),
                "unidades_PE":        filiais_un.get("PE","").strip(),
                "unidades_SE":        filiais_un.get("SE","").strip(),
                "positivacoes_geral": pos_geral.strip(),
                "positivacoes_AL":    filiais_pos.get("AL","").strip(),
                "positivacoes_PE":    filiais_pos.get("PE","").strip(),
                "positivacoes_SE":    filiais_pos.get("SE","").strip(),
                "valor_reposicao":    val_rep.strip(),
                "apurado_em":         datetime.now().strftime("%d/%m/%Y %H:%M"),
            }
            salvar_apuracao(ap_dados)
            st.success(f"✓ Apuração de **{id_ap}** salva com sucesso!")